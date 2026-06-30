"""
Genera un Excel de ejemplo para usar con crear_proceso_workiva.py
Uso: python generar_excel_ejemplo.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Proceso"

headers = [
    "titulo",
    "descripcion",
    "fecha_venc",
    "asignado_id",
    "aprobador_id",
    "modo_aprobacion",
    "fecha_aprobacion",
]

header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(color="FFFFFF", bold=True)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

ejemplos = [
    [
        "Revisión EE.FF. Q2 2026",
        "Revisar estados financieros del segundo trimestre",
        "15/07/2026",
        "user-id-001",
        "approver-id-001",
        "ONE",
        "20/07/2026",
    ],
    [
        "Conciliación cuentas por cobrar",
        "Conciliar saldos con SAP",
        "30/07/2026",
        "user-id-002",
        "approver-id-001;approver-id-002",
        "ALL",
        "05/08/2026",
    ],
    [
        "Validación nota revelaciones",
        "",
        "10/08/2026",
        "user-id-003",
        "",
        "ONE",
        "",
    ],
]

for fila in ejemplos:
    ws.append(fila)

col_widths = [35, 45, 15, 25, 40, 18, 18]
for col, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = w

nota = wb.create_sheet("Instrucciones")
instrucciones = [
    ["Campo",           "Descripción",                                          "Obligatorio"],
    ["titulo",          "Título de la tarea en Workiva",                        "Sí"],
    ["descripcion",     "Descripción detallada",                                "No"],
    ["fecha_venc",      "Fecha de vencimiento (DD/MM/YYYY)",                    "No"],
    ["asignado_id",     "ID del usuario Workiva responsable",                   "No"],
    ["aprobador_id",    "ID(s) del aprobador; separados por ';' si son varios", "No"],
    ["modo_aprobacion", "ONE = cualquiera aprueba | ALL = todos deben aprobar", "No (default: ONE)"],
    ["fecha_aprobacion","Fecha límite para la aprobación (DD/MM/YYYY)",         "No"],
    [],
    ["Nota:", "Para obtener los IDs de usuarios, consulte la API de Workiva o el panel de administración.", ""],
]
for fila in instrucciones:
    nota.append(fila)
nota.column_dimensions["A"].width = 20
nota.column_dimensions["B"].width = 60
nota.column_dimensions["C"].width = 20

path = "proceso_ejemplo.xlsx"
wb.save(path)
print(f"Archivo creado: {path}")
print("Edite los IDs de usuario y ejecute:")
print("  python crear_proceso_workiva.py proceso_ejemplo.xlsx")
