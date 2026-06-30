"""
Crea tareas en Workiva a partir de un archivo Excel.

Estructura esperada del Excel (hoja "Proceso"):
  Columna A: titulo         - Título de la tarea (obligatorio)
  Columna B: descripcion    - Descripción de la tarea
  Columna C: fecha_venc     - Fecha de vencimiento (DD/MM/YYYY o YYYY-MM-DD)
  Columna D: asignado_id    - ID de usuario Workiva asignado
  Columna E: aprobador_id   - ID(s) de usuario aprobador (separados por ";" si son varios)
  Columna F: modo_aprobacion- "ONE" (cualquiera aprueba) o "ALL" (todos deben aprobar), default "ONE"
  Columna G: fecha_aprobacion- Fecha límite de aprobación (DD/MM/YYYY o YYYY-MM-DD)

Uso:
  python crear_proceso_workiva.py proceso.xlsx
  python crear_proceso_workiva.py proceso.xlsx --hoja "Mi Hoja" --dry-run
"""

import sys
import os
import json
import argparse
import warnings
from datetime import datetime

import requests
from urllib3.exceptions import InsecureRequestWarning
from dotenv import load_dotenv

try:
    import openpyxl
except ImportError:
    print("Error: se requiere openpyxl. Instala con: pip install openpyxl")
    sys.exit(1)

load_dotenv()

warnings.filterwarnings("ignore", category=InsecureRequestWarning)
VERIFY_SSL = False

TOKEN_URL = "https://api.app.wdesk.com/iam/v1/oauth2/token"
API_BASE  = "https://api.app.wdesk.com"
API_VER   = "2026-01-01"

CLIENT_ID     = os.getenv("WORKIVA_CLIENT_ID")
CLIENT_SECRET = os.getenv("WORKIVA_CLIENT_SECRET")


# ── Auth ──────────────────────────────────────────────────────────────────────

def get_token() -> str:
    resp = requests.post(
        TOKEN_URL,
        json={
            "grant_type":    "client_credentials",
            "client_id":     CLIENT_ID,
            "client_secret": CLIENT_SECRET,
        },
        headers={"Content-Type": "application/json"},
        timeout=30,
        verify=VERIFY_SSL,
    )
    resp.raise_for_status()
    return resp.json()["access_token"]


def get_session() -> requests.Session:
    token = get_token()
    s = requests.Session()
    s.headers.update({
        "Authorization": f"Bearer {token}",
        "Content-Type":  "application/json",
        "Accept":        "application/json",
        "X-Version":     API_VER,
    })
    s.verify = VERIFY_SSL
    return s


# ── Excel ─────────────────────────────────────────────────────────────────────

def parse_date(value) -> str | None:
    """Convierte fecha a formato ISO-8601 (YYYY-MM-DDTHH:MM:SSZ)."""
    if not value:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%dT%H:%M:%SZ")
    s = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%dT00:00:00Z")
        except ValueError:
            continue
    print(f"  [WARN] Fecha no reconocida: '{s}' — se omite")
    return None


def leer_excel(path: str, hoja: str = "Proceso") -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if hoja not in wb.sheetnames:
        # intenta la primera hoja si no existe la indicada
        print(f"  [WARN] Hoja '{hoja}' no encontrada. Hojas disponibles: {wb.sheetnames}")
        hoja = wb.sheetnames[0]
        print(f"  [INFO] Usando hoja '{hoja}'")
    ws = wb[hoja]

    filas = list(ws.iter_rows(min_row=2, values_only=True))
    tareas = []
    for i, fila in enumerate(filas, start=2):
        if not any(fila):
            continue  # fila vacía

        titulo        = str(fila[0]).strip() if fila[0] else None
        descripcion   = str(fila[1]).strip() if len(fila) > 1 and fila[1] else ""
        fecha_venc    = parse_date(fila[2]) if len(fila) > 2 else None
        asignado_id   = str(fila[3]).strip() if len(fila) > 3 and fila[3] else None
        aprobadores   = str(fila[4]).strip() if len(fila) > 4 and fila[4] else None
        modo_aprobac  = str(fila[5]).strip().upper() if len(fila) > 5 and fila[5] else "ONE"
        fecha_aprob   = parse_date(fila[6]) if len(fila) > 6 else None

        if not titulo:
            print(f"  [WARN] Fila {i}: sin título — se omite")
            continue

        if modo_aprobac not in ("ONE", "ALL"):
            print(f"  [WARN] Fila {i}: modo_aprobacion '{modo_aprobac}' no válido, usando 'ONE'")
            modo_aprobac = "ONE"

        tarea: dict = {"title": titulo}
        if descripcion:
            tarea["description"] = descripcion
        if fecha_venc:
            tarea["dueDate"] = fecha_venc
        if asignado_id:
            tarea["assignees"] = [{"id": asignado_id, "type": "USER"}]
        if aprobadores:
            ids = [a.strip() for a in aprobadores.split(";") if a.strip()]
            paso = {
                "completionMode": modo_aprobac,
                "participants":   [{"id": uid, "type": "USER"} for uid in ids],
            }
            if fecha_aprob:
                paso["dueDate"] = fecha_aprob
            tarea["approvalSteps"] = [paso]

        tareas.append(tarea)

    return tareas


# ── API ───────────────────────────────────────────────────────────────────────

def crear_tarea(session: requests.Session, payload: dict, dry_run: bool = False) -> dict | None:
    if dry_run:
        print(f"  [DRY-RUN] POST /tasks → {json.dumps(payload, ensure_ascii=False)}")
        return {"id": "dry-run", "title": payload.get("title")}

    resp = session.post(f"{API_BASE}/tasks", json=payload, timeout=30)
    if resp.status_code == 201:
        return resp.json()
    print(f"  [ERROR] {resp.status_code}: {resp.text}")
    return None


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Crea tareas en Workiva desde un archivo Excel."
    )
    parser.add_argument("excel", help="Ruta al archivo Excel (.xlsx)")
    parser.add_argument("--hoja", default="Proceso", help="Nombre de la hoja (default: Proceso)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Solo muestra las tareas sin crearlas en Workiva")
    args = parser.parse_args()

    if not os.path.isfile(args.excel):
        print(f"Error: archivo '{args.excel}' no encontrado.")
        sys.exit(1)

    if not args.dry_run:
        if not CLIENT_ID or not CLIENT_SECRET:
            print("Error: WORKIVA_CLIENT_ID y WORKIVA_CLIENT_SECRET deben estar en .env")
            sys.exit(1)

    print(f"Leyendo '{args.excel}' (hoja: '{args.hoja}')...")
    tareas = leer_excel(args.excel, args.hoja)
    print(f"  → {len(tareas)} tarea(s) encontrada(s)\n")

    if not tareas:
        print("No hay tareas para crear.")
        return

    session = None
    if not args.dry_run:
        print("Conectando a Workiva...")
        session = get_session()
        print("  → Conexión exitosa\n")

    creadas = 0
    errores = 0
    for i, tarea in enumerate(tareas, 1):
        print(f"[{i}/{len(tareas)}] '{tarea['title']}'")
        resultado = crear_tarea(session, tarea, dry_run=args.dry_run)
        if resultado:
            tid = resultado.get("id", "?")
            print(f"  ✓ Creada — ID: {tid}")
            creadas += 1
        else:
            errores += 1

    print(f"\nResumen: {creadas} creada(s), {errores} error(es).")


if __name__ == "__main__":
    main()
