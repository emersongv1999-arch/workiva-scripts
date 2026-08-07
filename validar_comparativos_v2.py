#!/usr/bin/env python3
"""
validar_comparativos_v2.py  ←  ESPEJO que usa workiva_mcp_v2.py
================================================================
Idéntico a validar_comparativos.py pero carga workiva_mcp_v2.py
en lugar de workiva_mcp.py para aprovechar la detección de
columnas EERR (hoja C y equivalentes en Q2/Q3/Q4).

USO:
  python validar_comparativos_v2.py
      → pide interactivamente sociedad, año base, IND/CONSO y trimestre.

  python validar_comparativos_v2.py E110 2026 Q2
  python validar_comparativos_v2.py E200 2026 Q1 --tipo IND
  python validar_comparativos_v2.py --id <spreadsheet_id>

SALIDA: exit code 0 = todo calza, 2 = hay hallazgos, 1 = error.
"""

import argparse
import asyncio
import importlib.util
import json
import logging
import os
import re
import sys
from datetime import datetime
from pathlib import Path

logging.getLogger("httpx").setLevel(logging.WARNING)


# ── Cargar workiva_mcp_v2 desde la misma carpeta ─────────────────────────────

def _load_w():
    here     = Path(__file__).parent
    mcp_path = here / "workiva_mcp_v2.py"
    if not mcp_path.exists():
        print(f"ERROR: No se encuentra {mcp_path}")
        sys.exit(1)
    spec = importlib.util.spec_from_file_location("workiva_mcp_v2", mcp_path)
    mod  = importlib.util.module_from_spec(spec)
    sys.modules["workiva_mcp_v2"] = mod
    spec.loader.exec_module(mod)
    return mod


w = _load_w()

MES_POR_TRIMESTRE = {
    "Q1": "03", "Q2": "06", "Q3": "09", "Q4": "12",
    "1": "03", "2": "06", "3": "09", "4": "12",
    "03": "03", "06": "06", "09": "09", "12": "12",
}


def _pedir(texto: str, validar, default: str | None = None) -> str:
    while True:
        v = input(texto)
        for bom in (chr(0xFEFF), chr(0xEF) + chr(0xBB) + chr(0xBF)):
            v = v.removeprefix(bom)
        v = v.strip()
        if not v and default is not None:
            return default
        if validar(v):
            return v
        print("   valor no válido, intenta de nuevo")


def pedir_opciones() -> tuple[str, str, str, str]:
    print("=== Validador de comparativos Workiva V2 (cierre normal) ===\n")
    sociedad = _pedir("Sociedad (ej E110, E200): ",
                      lambda v: re.fullmatch(r"[Ee]\d+", v) is not None).upper()
    anio = _pedir("Año base (ej 2026): ",
                  lambda v: re.fullmatch(r"\d{4}", v) is not None)
    tipo_in = _pedir("Tipo de cierre [1] CONSO  [2] IND  (Enter = CONSO): ",
                     lambda v: v.upper() in ("1", "2", "CONSO", "IND"),
                     default="CONSO")
    tipo = "IND" if tipo_in.upper() in ("2", "IND") else "CONSO"
    trimestre = _pedir("Trimestre (Q1/Q2/Q3/Q4, 1-4 o mes 03/06/09/12): ",
                       lambda v: v.upper() in MES_POR_TRIMESTRE)
    print()
    return sociedad, anio, tipo, trimestre


async def resolver_spreadsheet(sociedad: str, anio: str, trimestre: str,
                               tipo: str) -> tuple[str, str] | None:
    mm = MES_POR_TRIMESTRE[trimestre.upper()]
    all_files = await w._load_all_files()
    patron    = re.compile(rf"^{re.escape(sociedad)}_{tipo}_{mm}[-_]{anio}_")
    matches   = {n: i for n, i in all_files.items() if patron.match(n)}

    if len(matches) == 1:
        name, ss_id = next(iter(matches.items()))
        print(f"Archivo encontrado: {name}")
        return ss_id, name

    if len(matches) > 1:
        print("ERROR: más de un archivo calza; usa --id para elegir uno:")
        for n, i in sorted(matches.items()):
            print(f"  {n}  (id {i})")
        return None

    print(f"ERROR: no se encontró archivo para {sociedad} {tipo} {mm}-{anio}.")
    disponibles = sorted(
        n for n in all_files
        if n.startswith(f"{sociedad}_") and "Base Notas" in n
    )
    if disponibles:
        print("Archivos Base Notas disponibles para esa sociedad:")
        for n in disponibles:
            print(f"  {n}")
    return None


# ── Reglas de alcance por hoja ───────────────────────────────────────────────
#
# 1) LIMITE_FILAS: última fila válida de cada nota. Más abajo de esa fila las
#    hojas traen bloques auxiliares (PPT, cuadros de apoyo, etc.) que no son
#    comparativo y generaban hallazgos falsos.
#
# 2) Subhojas: en los archivos CONSOLIDADOS cada nota se repite una o más
#    veces como hoja de apoyo ("27.- CGEM Conso", "73.- PPA", etc.). Son
#    apoyo de la nota real y NO deben validarse.
#
#    La detección NO usa una lista de nombres: esos nombres cambian de una
#    sociedad a otra. Se usa el prefijo de la nota ("27", "F", "A"), que es
#    su identificador: la nota real es la PRIMERA hoja del archivo con ese
#    prefijo y cualquier hoja posterior que repita el prefijo es subhoja.
#
#    Consecuencia importante: una hoja con prefijo único jamás se descarta.

RANGO_FILAS: dict[str, tuple[int | None, int | None]] = {
    "A":   (None, 43),
    "B":   (None, 55),
    "C":   (None, 54),
    "D":   (None, 77),
    "K":   (None, 14),
    "13":  (None, 14),
    "14":  (None, 45),
    "15":  (None, 14),
    "17":  (None, 15),
    "22":  (7,    11),
    "53":  (None, 25),
    "55":  (7,    18),
    "57":  (None, 41),
    "74":  (None, 22),
    "77":  (None, 40),
    "85":  (None, 37),
    "90":  (None, 25),
    "95":  (None, 16),
    "104": (None, 22),
    "105": (None, 11),
    "106": (None, 12),
    "107": (None, 20),
    "108": (None, 13),
    "109": (None, 19),
    "110": (None, 24),
    "111": (None, 37),
    "113": (None, 44),
    "114": (None, 29),
    "118": (None, 21),
    "120": (None, 43),
    "121": (None, 53),
}

# "A.- Activos" -> "A" ; "27. CGEM Conso" -> "27" ; "55.-CGE" -> "55"
_RE_PREFIJO = re.compile(r"^\s*([A-Za-zÑñ0-9]{1,4})\s*\.\s*-?\s*\S")


def prefijo_hoja(nombre: str) -> str:
    """Prefijo de la nota ("A", "27"), o "" si el nombre no lo lleva."""
    m = _RE_PREFIJO.match(nombre or "")
    return m.group(1).upper() if m else ""


class DetectorSubhojas:
    """Marca como subhoja toda hoja que repita el prefijo de una anterior.

    Se alimenta en el mismo orden en que Workiva devuelve las hojas, que es
    el orden del archivo: la nota real va primero y sus hojas de apoyo la
    siguen. Una hoja sin prefijo, o con un prefijo que aparece una sola vez,
    nunca se marca.
    """

    def __init__(self) -> None:
        self._vistos: dict[str, str] = {}

    def es_subhoja(self, nombre: str) -> str | None:
        """Devuelve el nombre de la nota real si `nombre` es subhoja suya."""
        prefijo = prefijo_hoja(nombre)
        if not prefijo:
            return None
        padre = self._vistos.get(prefijo)
        if padre is None:
            self._vistos[prefijo] = nombre
            return None
        return padre


def rango_filas(nombre: str) -> tuple[int | None, int | None]:
    """Devuelve (desde, hasta) para esa hoja. None = sin límite en ese extremo."""
    return RANGO_FILAS.get(prefijo_hoja(nombre), (None, None))


def _nombre_pestana(nombre: str, usados: set[str]) -> str:
    limpio = re.sub(r"[\[\]:*?/\\]", " ", nombre)[:31].rstrip()
    base, n = limpio, 2
    while limpio in usados:
        sufijo = f" ({n})"
        limpio = base[: 31 - len(sufijo)] + sufijo
        n += 1
    usados.add(limpio)
    return limpio


def exportar_excel(ruta: str, titulo: str, subtitulo: str,
                   hojas: list[dict]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    bold = Font(bold=True)
    wb   = Workbook()

    ws = wb.active
    ws.title = "Resumen"
    ws.cell(row=1, column=1, value=titulo).font = bold
    ws.cell(row=2, column=1, value=subtitulo).font = bold
    encabezados = ["Nota", "Hoja Excel", "Filas", "OK", "Hallazgo", "No procesado",
                   "Sin correspondencia"]
    for j, h in enumerate(encabezados, start=1):
        ws.cell(row=4, column=j, value=h).font = bold
    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 34
    ws.freeze_panes = "A5"

    usados: set[str] = {"Resumen"}
    for i, hoja in enumerate(hojas, start=5):
        filas = hoja["filas"]
        tab   = _nombre_pestana(hoja["nombre"], usados) if filas else "-"
        ws.cell(row=i, column=1, value=hoja["nombre"])
        ws.cell(row=i, column=2, value=tab)
        ws.cell(row=i, column=3, value=len(filas))
        ws.cell(row=i, column=4, value=sum(1 for f in filas if f["estado"] == "OK"))
        ws.cell(row=i, column=5, value=sum(1 for f in filas if f["estado"] == "HALLAZGO"))
        ws.cell(row=i, column=6, value=sum(1 for f in filas if f["estado"] == "NO PROCESADO"))
        ws.cell(row=i, column=7,
                value=sum(1 for f in filas if f["estado"] == "SIN CORRESPONDENCIA"))

        if not filas:
            continue
        ws_n = wb.create_sheet(tab)
        ws_n.cell(row=1, column=1, value=hoja["nombre"]).font = bold
        cols = ["Fila", "Etiqueta", "Valor declarado (comparativo)",
                "Valor real (fuente)", "Estado", "Nota"]
        for j, h in enumerate(cols, start=1):
            ws_n.cell(row=2, column=j, value=h).font = bold
        for r, f in enumerate(sorted(filas, key=lambda x: (x["fila"], x["etiqueta"])),
                              start=3):
            ws_n.cell(row=r, column=1, value=f["fila"])
            ws_n.cell(row=r, column=2, value=f["etiqueta"])
            c3 = ws_n.cell(row=r, column=3, value=f["destino"])
            c4 = ws_n.cell(row=r, column=4, value=f["fuente"])
            c3.number_format = "#,##0"
            c4.number_format = "#,##0"
            ws_n.cell(row=r, column=5, value=f["estado"])
            ws_n.cell(row=r, column=6, value=f["nota"])
        ws_n.column_dimensions["B"].width = 60
        ws_n.column_dimensions["C"].width = 22
        ws_n.column_dimensions["D"].width = 22
        ws_n.column_dimensions["E"].width = 14
        ws_n.column_dimensions["F"].width = 30
        ws_n.freeze_panes = "A3"

    wb.save(ruta)


async def validar(spreadsheet_id: str, etiqueta: str, max_sheets: int = 50) -> int:
    offset = 0
    total_equal = total_diff = total_sin_corr = 0
    candidatas = "?"
    info: dict = {}
    hojas: list[dict] = []
    encabezado_impreso = False
    omitidas_subhoja = 0
    filas_fuera      = 0
    detector         = DetectorSubhojas()

    while True:
        raw = await w.workiva_fill_comparatives(
            w.FillComparativesInput(
                spreadsheet_id=spreadsheet_id,
                dry_run=True,
                sheet_offset=offset,
                max_sheets=max_sheets,
                detalle_filas=True,
            )
        )
        try:
            r = json.loads(raw)
        except json.JSONDecodeError:
            print(f"ERROR del conector: {raw}")
            return 1

        if "warning" in r:
            print(f"ADVERTENCIA: {r['warning']}")
            return 1

        if not encabezado_impreso:
            candidatas = r.get("total_candidate_sheets", "?")
            info = {
                "current_end":    r["current_end"],
                "prior_end":      r["prior_end"],
                "source_balance": r.get("source_balance", "?"),
                "source_eerr":    r.get("source_eerr", "No encontrado"),
            }
            print(f"Archivo destino   : {etiqueta}")
            print(f"Período actual    : {r['current_end']}")
            print(f"Comparativo bal.  : {r['prior_end']}")
            print(f"Fuente balance    : {info['source_balance']}")
            print(f"Fuente EERR       : {info['source_eerr']}")
            print(f"Hojas a validar   : {candidatas}"
                  f" (excluidos {r.get('skipped_desglose_sociedad', 0)} desgloses por sociedad)")
            print("-" * 70)
            encabezado_impreso = True

        for sh in r.get("sheets_processed", []):
            nombre_hoja = sh["sheet"]

            # Subhoja de apoyo de un consolidado: no se valida.
            padre = detector.es_subhoja(nombre_hoja)
            if padre is not None:
                omitidas_subhoja += 1
                continue

            desde, hasta = rango_filas(nombre_hoja)
            filas_hoja: list[dict] = []
            comps = sh.get("comparacion", [])
            debug_cols = {}
            for _dbg in sh.get("_debug_src_cols", []):
                _letra = _dbg.split("(", 1)[0]
                debug_cols[_letra] = _dbg
            for comp in comps:
                contexto = (comp.get("contexto") or "").strip()
                tipo_col = comp.get("tipo", "bal")
                for f in comp.get("filas", []):
                    # Fuera del alcance de la nota (bloques auxiliares al pie).
                    if (desde is not None and f["fila"] < desde) or (hasta is not None and f["fila"] > hasta):
                        filas_fuera += 1
                        continue
                    if f["estado"] == "OK":
                        total_equal += 1
                    elif f["estado"] == "SIN CORRESPONDENCIA":
                        total_sin_corr += 1
                    else:
                        total_diff += 1
                    base = f["etiqueta"] or f"(fila {f['fila']})"
                    if contexto:
                        etiq = f"{base} - {contexto}"
                    elif len(comps) > 1:
                        etiq = f"{base} (col {comp['col']} {tipo_col})"
                    else:
                        etiq = base
                    if f["estado"] == "HALLAZGO":
                        try:
                            nota = f"difiere en {float(f['destino']) - float(f['fuente']):,.0f}"
                        except (TypeError, ValueError):
                            nota = "difiere"
                        _dbg = debug_cols.get(comp["col"])
                        if _dbg:
                            nota = f"{nota}  |  {_dbg}"
                    elif f["estado"] == "NO PROCESADO":
                        nota = "valor destino no numérico"
                    elif f["estado"] == "SIN CORRESPONDENCIA":
                        nota = "la fila no existe en el archivo fuente: revisar manualmente"
                    else:
                        nota = None
                    filas_hoja.append({
                        "fila": f["fila"], "etiqueta": etiq,
                        "destino": f["destino"], "fuente": f["fuente"],
                        "estado": f["estado"], "nota": nota,
                    })
            hojas.append({"nombre": nombre_hoja, "filas": filas_hoja})

        print(f"  lote offset {r['sheet_offset']:>3}: {r['batch_size']} hojas revisadas")

        if not r.get("has_more"):
            break
        offset = r["next_offset"]

    print("-" * 70)
    print(f"RESUMEN: {len(hojas)} hojas | "
          f"{total_equal} valores iguales | {total_diff} con hallazgo/no procesado")
    if omitidas_subhoja:
        print(f"  ({omitidas_subhoja} subhoja(s) de apoyo omitidas, ver detalle arriba)")
    if filas_fuera:
        print(f"  ({filas_fuera} fila(s) fuera del alcance de su nota, omitidas)")
    if total_sin_corr:
        print(f"  ({total_sin_corr} fila(s) sin correspondencia en el archivo fuente, "
              f"revisar manualmente en el Excel)")

    m = re.match(r"(E\d+)_(IND|CONSO)_(\d{2})[-_](\d{4})", etiqueta)
    if m:
        base      = f"{m.group(1)}_{m.group(2)}_{m.group(3)}-{m.group(4)}"
        subtitulo = f"{m.group(1)} {m.group(2)} {m.group(3)}-{m.group(4)}"
    else:
        base      = re.sub(r'[\\/:*?"<>|\s]+', "_", etiqueta)
        subtitulo = etiqueta
    subtitulo += (f" — bal {info.get('prior_end','?')}"
                  f" / EERR {info.get('source_eerr','?')}"
                  f" — revisado {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    ruta = os.path.abspath(f"detalle_filas_{base}.xlsx")
    exportar_excel(
        ruta,
        "Detalle fila por fila — comparativo declarado vs. archivo fuente (V2)",
        subtitulo,
        hojas,
    )
    print(f"\nExcel generado: {ruta}")

    return 2 if total_diff else 0


async def main() -> int:
    parser = argparse.ArgumentParser(
        description="Valida comparativos (balance + EERR) de un Base Notas de Workiva.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("sociedad",   nargs="?")
    parser.add_argument("anio",       nargs="?")
    parser.add_argument("trimestre",  nargs="?")
    parser.add_argument("--tipo",     choices=["CONSO", "IND"], default="CONSO")
    parser.add_argument("--id",       dest="spreadsheet_id")
    parser.add_argument("--lote",     type=int, default=50)
    args = parser.parse_args()

    if args.spreadsheet_id:
        ss_id, etiqueta = args.spreadsheet_id, args.spreadsheet_id
    else:
        if args.sociedad and args.anio and args.trimestre:
            sociedad, anio, tipo, trimestre = (
                args.sociedad.upper(), args.anio, args.tipo, args.trimestre
            )
            if trimestre.upper() not in MES_POR_TRIMESTRE:
                print(f"ERROR: trimestre '{trimestre}' no válido.")
                return 1
        else:
            sociedad, anio, tipo, trimestre = pedir_opciones()

        encontrado = await resolver_spreadsheet(sociedad, anio, trimestre, tipo)
        if not encontrado:
            return 1
        ss_id, etiqueta = encontrado

    return await validar(ss_id, etiqueta, args.lote)


if __name__ == "__main__":
    try:
        sys.exit(asyncio.run(main()))
    except KeyboardInterrupt:
        print("\nCancelado.")
        sys.exit(1)
