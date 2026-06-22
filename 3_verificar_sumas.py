"""
PASO 3 — Verificar sumas en tablas de los archivos .docx descargados.
Lee todos los .docx de ./docx_junio_2026/ y compara totales declarados
con la suma calculada de sus filas.

Uso:
    python 3_verificar_sumas.py

Genera reporte en: reporte_sumas_junio_2026.xlsx y reporte_sumas_junio_2026.txt
"""
import os
import re
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")

try:
    from docx import Document
except ImportError:
    print("Falta librería: pip install python-docx")
    raise

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False
    print("Aviso: openpyxl no instalado. Solo se generará reporte .txt")

DOCX_DIR     = "docx_junio_2026"
TXT_REPORT   = "reporte_sumas_junio_2026.txt"
XLSX_REPORT  = "reporte_sumas_junio_2026.xlsx"

# Patrones que suelen indicar una fila de total/subtotal en los EE.FF.
TOTAL_KEYWORDS = [
    "total", "subtotal", "suma", "activos", "pasivos", "patrimonio",
    "resultado", "ganancia", "utilidad", "pérdida", "perdida",
    "flujo", "efectivo", "activo total", "pasivo total",
]


def limpiar_numero(texto: str):
    """Convierte un texto con formato numérico chileno a float. Retorna None si no es número."""
    t = texto.strip().replace("\xa0", "").replace(" ", "")
    if not t or t in ("-", "—", "–"):
        return None
    # Eliminar paréntesis (números negativos en EEFF)
    negativo = t.startswith("(") and t.endswith(")")
    t = t.strip("()")
    # Formato chileno: punto como separador de miles, coma como decimal
    t = t.replace(".", "").replace(",", ".")
    try:
        valor = float(t)
        return -valor if negativo else valor
    except ValueError:
        return None


def es_fila_total(celdas: list[str]) -> bool:
    """Heurística: la primera celda contiene alguna palabra clave de total."""
    primera = celdas[0].lower() if celdas else ""
    return any(kw in primera for kw in TOTAL_KEYWORDS)


def analizar_tabla(tabla, tabla_idx: int, nombre_doc: str) -> list[dict]:
    """
    Analiza una tabla buscando filas de totales y comparándolas con la suma
    de las filas numéricas anteriores en cada columna.
    Retorna lista de discrepancias o confirmaciones.
    """
    resultados = []
    filas = tabla.rows

    # Extraer datos: lista de listas de strings
    datos = []
    for fila in filas:
        fila_texto = [c.text.strip() for c in fila.cells]
        datos.append(fila_texto)

    if not datos:
        return resultados

    num_cols = max(len(f) for f in datos)

    # Acumular sumas por columna y detectar filas total
    acumulado = [0.0] * num_cols
    tiene_datos = [False] * num_cols

    for i, fila in enumerate(datos):
        # Rellenar celdas faltantes
        while len(fila) < num_cols:
            fila.append("")

        if es_fila_total(fila):
            # Comparar cada columna numérica con el acumulado
            for col in range(1, num_cols):
                declarado = limpiar_numero(fila[col])
                if declarado is None or not tiene_datos[col]:
                    continue
                diff = abs(declarado - acumulado[col])
                ok = diff < 1  # tolerancia 1 unidad (miles de CLP)
                resultados.append({
                    "documento":  nombre_doc,
                    "tabla":      tabla_idx + 1,
                    "fila":       i + 1,
                    "descripcion": fila[0][:60],
                    "columna":    col + 1,
                    "declarado":  declarado,
                    "calculado":  round(acumulado[col], 2),
                    "diferencia": round(declarado - acumulado[col], 2),
                    "ok":         ok,
                })
            # Resetear acumulado tras encontrar un total
            acumulado = [0.0] * num_cols
            tiene_datos = [False] * num_cols
        else:
            # Sumar valores numéricos
            for col in range(1, num_cols):
                v = limpiar_numero(fila[col]) if col < len(fila) else None
                if v is not None:
                    acumulado[col] += v
                    tiene_datos[col] = True

    return resultados


def procesar_docx(ruta: Path) -> list[dict]:
    resultados = []
    try:
        doc = Document(ruta)
        for idx, tabla in enumerate(doc.tables):
            res = analizar_tabla(tabla, idx, ruta.name)
            resultados.extend(res)
    except Exception as e:
        resultados.append({
            "documento": ruta.name,
            "tabla": 0,
            "fila": 0,
            "descripcion": f"ERROR al leer archivo: {e}",
            "columna": 0,
            "declarado": None,
            "calculado": None,
            "diferencia": None,
            "ok": False,
        })
    return resultados


def generar_txt(todos: list[dict]):
    errores  = [r for r in todos if not r["ok"]]
    correctos = [r for r in todos if r["ok"]]

    with open(TXT_REPORT, "w", encoding="utf-8") as f:
        f.write("=" * 80 + "\n")
        f.write("REPORTE DE VERIFICACIÓN DE SUMAS — JUNIO 2026\n")
        f.write("=" * 80 + "\n\n")
        f.write(f"Total verificaciones: {len(todos)}\n")
        f.write(f"Correctas:            {len(correctos)}\n")
        f.write(f"Con diferencia:       {len(errores)}\n\n")

        if errores:
            f.write("=" * 80 + "\n")
            f.write("DIFERENCIAS ENCONTRADAS\n")
            f.write("=" * 80 + "\n")
            for r in errores:
                if r["diferencia"] is None:
                    f.write(f"\n[ERROR] {r['documento']}\n  {r['descripcion']}\n")
                    continue
                f.write(f"\nDocumento: {r['documento']}\n")
                f.write(f"  Tabla {r['tabla']}, Fila {r['fila']}, Col {r['columna']}: {r['descripcion']}\n")
                f.write(f"  Declarado:  {r['declarado']:>15,.2f}\n")
                f.write(f"  Calculado:  {r['calculado']:>15,.2f}\n")
                f.write(f"  Diferencia: {r['diferencia']:>15,.2f}  ◄ REVISAR\n")
        else:
            f.write("No se encontraron diferencias.\n")

    print(f"Reporte TXT: {TXT_REPORT}")


def generar_xlsx(todos: list[dict]):
    if not EXCEL_OK:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Verificacion Sumas"

    # Estilos
    header_fill  = PatternFill("solid", fgColor="1F4E79")
    header_font  = Font(bold=True, color="FFFFFF")
    ok_fill      = PatternFill("solid", fgColor="C6EFCE")
    err_fill     = PatternFill("solid", fgColor="FFC7CE")
    center       = Alignment(horizontal="center")

    headers = ["Documento", "Tabla", "Fila", "Descripción", "Col",
               "Declarado", "Calculado", "Diferencia", "Estado"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_font and header_fill
        cell.font = header_font
        cell.alignment = center

    for r in todos:
        estado = "OK" if r["ok"] else "REVISAR"
        ws.append([
            r["documento"],
            r["tabla"],
            r["fila"],
            r["descripcion"],
            r["columna"],
            r["declarado"],
            r["calculado"],
            r["diferencia"],
            estado,
        ])
        fill = ok_fill if r["ok"] else err_fill
        for cell in ws[ws.max_row]:
            cell.fill = fill

    # Anchos de columna
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["D"].width = 40
    for col in ["B", "C", "E"]:
        ws.column_dimensions[col].width = 8
    for col in ["F", "G", "H"]:
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["I"].width = 10

    # Hoja resumen
    ws2 = wb.create_sheet("Resumen por Documento")
    ws2.append(["Documento", "Verificaciones", "OK", "Con diferencia"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font

    docs = {}
    for r in todos:
        d = r["documento"]
        if d not in docs:
            docs[d] = {"total": 0, "ok": 0, "err": 0}
        docs[d]["total"] += 1
        if r["ok"]:
            docs[d]["ok"] += 1
        else:
            docs[d]["err"] += 1

    for nombre, cnt in sorted(docs.items()):
        ws2.append([nombre, cnt["total"], cnt["ok"], cnt["err"]])
        fill = ok_fill if cnt["err"] == 0 else err_fill
        for cell in ws2[ws2.max_row]:
            cell.fill = fill
    ws2.column_dimensions["A"].width = 55

    wb.save(XLSX_REPORT)
    print(f"Reporte Excel: {XLSX_REPORT}")


def main():
    docx_dir = Path(DOCX_DIR)
    if not docx_dir.exists():
        print(f"No existe la carpeta {DOCX_DIR}. Ejecuta primero: python 2_descargar_docx.py")
        return

    archivos = sorted(docx_dir.glob("*.docx"))
    if not archivos:
        print(f"No hay archivos .docx en {DOCX_DIR}")
        return

    print(f"Analizando {len(archivos)} archivos .docx...\n")
    todos = []
    for i, ruta in enumerate(archivos, 1):
        print(f"[{i}/{len(archivos)}] {ruta.name}")
        res = procesar_docx(ruta)
        ok  = sum(1 for r in res if r["ok"])
        err = len(res) - ok
        print(f"  Verificaciones: {len(res)} | OK: {ok} | Diferencias: {err}")
        todos.extend(res)

    total_err = sum(1 for r in todos if not r["ok"])
    print(f"\n{'='*60}")
    print(f"TOTAL verificaciones: {len(todos)}")
    print(f"Correctas:            {len(todos) - total_err}")
    print(f"Con diferencia:       {total_err}")
    print(f"{'='*60}")

    generar_txt(todos)
    generar_xlsx(todos)
    print("\nListo.")


if __name__ == "__main__":
    main()
