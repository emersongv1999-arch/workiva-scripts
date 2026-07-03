#!/usr/bin/env python3
"""
validar_comparativos.py
=======================
Valida los saldos comparativos de un archivo "Base Notas" de Workiva contra
su archivo fuente (el cierre de diciembre del año anterior), SIN escribir nada.

Solo trabaja con archivos del cierre normal (excluye las versiones (LC) y (CHN)).

Al terminar deja un Excel (detalle_filas_<sociedad>_<tipo>_<MM-AAAA>.xlsx en la
carpeta actual) con el detalle fila por fila:
  - Hoja "Resumen": índice con Nota / Hoja Excel / Filas / OK / Hallazgo / No procesado
  - Una hoja por nota revisada: Fila / Etiqueta / Valor declarado (comparativo) /
    Valor real (fuente) / Estado / Nota

USO:
  python validar_comparativos.py
      → pide interactivamente sociedad, año base, IND/CONSO y trimestre.

  También acepta los datos como argumentos para saltarse las preguntas:
      python validar_comparativos.py E110 2026 Q2
      python validar_comparativos.py E200 2026 Q1 --tipo IND
      python validar_comparativos.py --id <spreadsheet_id>

SALIDA: exit code 0 = todo calza, 2 = hay hallazgos, 1 = error.
"""

import argparse
import asyncio
import json
import logging
import os
import re
import sys
from datetime import datetime

logging.getLogger("httpx").setLevel(logging.WARNING)  # silenciar log por request

import workiva_mcp as w

MES_POR_TRIMESTRE = {
    "Q1": "03", "Q2": "06", "Q3": "09", "Q4": "12",
    "1": "03", "2": "06", "3": "09", "4": "12",
    "03": "03", "06": "06", "09": "09", "12": "12",
}


def _pedir(texto: str, validar, default: str | None = None) -> str:
    """Pide un valor por consola hasta que sea válido. Enter usa el default."""
    while True:
        v = input(texto)
        # Tolerar BOM si la entrada viene por pipe (U+FEFF o sus bytes UTF-8)
        for bom in (chr(0xFEFF), chr(0xEF) + chr(0xBB) + chr(0xBF)):
            v = v.removeprefix(bom)
        v = v.strip()
        if not v and default is not None:
            return default
        if validar(v):
            return v
        print("   valor no válido, intenta de nuevo")


def pedir_opciones() -> tuple[str, str, str, str]:
    """Menú interactivo: sociedad, año base, tipo y trimestre."""
    print("=== Validador de comparativos Workiva (cierre normal) ===\n")
    sociedad = _pedir("Sociedad (ej E110, E200): ",
                      lambda v: re.fullmatch(r"[Ee]\d+", v) is not None).upper()
    anio = _pedir("Año base (ej 2026): ",
                  lambda v: re.fullmatch(r"\d{4}", v) is not None)
    tipo_in = _pedir("Tipo de cierre [1] CONSO  [2] IND  (Enter = CONSO): ",
                     lambda v: v.upper() in ("1", "2", "CONSO", "IND"),
                     default="CONSO")
    tipo = "IND" if tipo_in.upper() in ("2", "IND") else "CONSO"
    trimestre = _pedir("Trimestre (Q1/Q2/Q3/Q4, 1-4 o mes 03/06/09/12): ",
                       lambda v: v.upper() in MES_POR_TRIMESTRE)
    print()
    return sociedad, anio, tipo, trimestre


async def resolver_spreadsheet(sociedad: str, anio: str, trimestre: str,
                               tipo: str) -> tuple[str, str] | None:
    """Busca el (id, nombre) del cierre normal (sin prefijo LC/CHN)."""
    mm = MES_POR_TRIMESTRE[trimestre.upper()]

    all_files = await w._load_all_files()
    patron = re.compile(rf"^{re.escape(sociedad)}_{tipo}_{mm}[-_]{anio}_")
    matches = {n: i for n, i in all_files.items() if patron.match(n)}

    if len(matches) == 1:
        name, ss_id = next(iter(matches.items()))
        print(f"Archivo encontrado: {name}")
        return ss_id, name

    if len(matches) > 1:
        print("ERROR: más de un archivo calza; usa --id para elegir uno:")
        for n, i in sorted(matches.items()):
            print(f"  {n}  (id {i})")
        return None

    print(f"ERROR: no se encontró archivo para {sociedad} {tipo} {mm}-{anio}.")
    disponibles = sorted(
        n for n in all_files
        if n.startswith(f"{sociedad}_") and "Base Notas" in n
    )
    if disponibles:
        print("Archivos Base Notas disponibles para esa sociedad (cierre normal):")
        for n in disponibles:
            print(f"  {n}")
    return None


def _nombre_pestana(nombre: str, usados: set[str]) -> str:
    """Nombre válido y único para una pestaña de Excel (máx 31 chars)."""
    limpio = re.sub(r"[\[\]:*?/\\]", " ", nombre)[:31].rstrip()
    base, n = limpio, 2
    while limpio in usados:
        sufijo = f" ({n})"
        limpio = base[: 31 - len(sufijo)] + sufijo
        n += 1
    usados.add(limpio)
    return limpio


def exportar_excel(ruta: str, titulo: str, subtitulo: str,
                   hojas: list[dict]) -> None:
    """Escribe el Excel formato 'detalle fila por fila' (una pestaña por nota)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    bold = Font(bold=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "Resumen"
    ws.cell(row=1, column=1, value=titulo).font = bold
    ws.cell(row=2, column=1, value=subtitulo).font = bold
    encabezados = ["Nota", "Hoja Excel", "Filas", "OK", "Hallazgo", "No procesado"]
    for j, h in enumerate(encabezados, start=1):
        ws.cell(row=4, column=j, value=h).font = bold
    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 34
    ws.freeze_panes = "A5"

    usados: set[str] = {"Resumen"}
    for i, hoja in enumerate(hojas, start=5):
        filas = hoja["filas"]
        tab = _nombre_pestana(hoja["nombre"], usados) if filas else "-"
        ws.cell(row=i, column=1, value=hoja["nombre"])
        ws.cell(row=i, column=2, value=tab)
        ws.cell(row=i, column=3, value=len(filas))
        ws.cell(row=i, column=4, value=sum(1 for f in filas if f["estado"] == "OK"))
        ws.cell(row=i, column=5, value=sum(1 for f in filas if f["estado"] == "HALLAZGO"))
        ws.cell(row=i, column=6, value=sum(1 for f in filas if f["estado"] == "NO PROCESADO"))

        if not filas:
            continue
        ws_n = wb.create_sheet(tab)
        ws_n.cell(row=1, column=1, value=hoja["nombre"]).font = bold
        cols = ["Fila", "Etiqueta", "Valor declarado (comparativo)",
                "Valor real (fuente)", "Estado", "Nota"]
        for j, h in enumerate(cols, start=1):
            ws_n.cell(row=2, column=j, value=h).font = bold
        for r, f in enumerate(sorted(filas, key=lambda x: (x["fila"], x["etiqueta"])),
                              start=3):
            ws_n.cell(row=r, column=1, value=f["fila"])
            ws_n.cell(row=r, column=2, value=f["etiqueta"])
            c3 = ws_n.cell(row=r, column=3, value=f["destino"])
            c4 = ws_n.cell(row=r, column=4, value=f["fuente"])
            c3.number_format = "#,##0"
            c4.number_format = "#,##0"
            ws_n.cell(row=r, column=5, value=f["estado"])
            ws_n.cell(row=r, column=6, value=f["nota"])
        ws_n.column_dimensions["B"].width = 60
        ws_n.column_dimensions["C"].width = 22
        ws_n.column_dimensions["D"].width = 22
        ws_n.column_dimensions["E"].width = 14
        ws_n.column_dimensions["F"].width = 30
        ws_n.freeze_panes = "A3"

    wb.save(ruta)


async def validar(spreadsheet_id: str, etiqueta: str, max_sheets: int = 50) -> int:
    offset = 0
    total_equal = total_diff = 0
    candidatas = "?"
    info: dict = {}
    hojas: list[dict] = []          # [{nombre, filas: [{fila, etiqueta, destino, fuente, estado, nota}]}]
    encabezado_impreso = False

    while True:
        raw = await w.workiva_fill_comparatives(
            w.FillComparativesInput(
                spreadsheet_id=spreadsheet_id,
                dry_run=True,
                sheet_offset=offset,
                max_sheets=max_sheets,
                detalle_filas=True,
            )
        )
        try:
            r = json.loads(raw)
        except json.JSONDecodeError:
            print(f"ERROR del conector: {raw}")
            return 1

        if "warning" in r:
            print(f"ADVERTENCIA: {r['warning']}")
            return 1

        if not encabezado_impreso:
            candidatas = r.get("total_candidate_sheets", "?")
            info = {
                "current_end": r["current_end"],
                "prior_end": r["prior_end"],
                "source_balance": r.get("source_balance", "?"),
            }
            print(f"Archivo destino : {etiqueta}")
            print(f"Período actual  : {r['current_end']}")
            print(f"Comparativo     : {r['prior_end']}")
            print(f"Archivo fuente  : {info['source_balance']}")
            print(f"Hojas a validar : {candidatas}"
                  f" (excluidos {r.get('skipped_desglose_sociedad', 0)} desgloses por sociedad)")
            print("-" * 70)
            encabezado_impreso = True

        for sh in r.get("sheets_processed", []):
            filas_hoja: list[dict] = []
            comps = sh.get("comparacion", [])
            for comp in comps:
                total_equal += comp["iguales"]
                total_diff += comp["distintos"]
                contexto = (comp.get("contexto") or "").strip()
                for f in comp.get("filas", []):
                    base = f["etiqueta"] or f"(fila {f['fila']})"
                    if contexto:
                        etiq = f"{base} - {contexto}"
                    elif len(comps) > 1:
                        etiq = f"{base} (col {comp['col']})"
                    else:
                        etiq = base
                    if f["estado"] == "HALLAZGO":
                        try:
                            nota = f"difiere en {float(f['destino']) - float(f['fuente']):,.0f}"
                        except (TypeError, ValueError):
                            nota = "difiere"
                    elif f["estado"] == "NO PROCESADO":
                        nota = "valor destino no numérico"
                    else:
                        nota = None
                    filas_hoja.append({
                        "fila": f["fila"], "etiqueta": etiq,
                        "destino": f["destino"], "fuente": f["fuente"],
                        "estado": f["estado"], "nota": nota,
                    })
            # Incluir también hojas con columnas detectadas pero sin valores
            # comparables (todo fórmulas): aparecen en el índice con 0 filas.
            hojas.append({"nombre": sh["sheet"], "filas": filas_hoja})

        print(f"  lote offset {r['sheet_offset']:>3}: {r['batch_size']} hojas revisadas")

        if not r.get("has_more"):
            break
        offset = r["next_offset"]

    print("-" * 70)
    print(f"RESUMEN: {len(hojas)} hojas con columnas comparativas | "
          f"{total_equal} valores iguales | {total_diff} con hallazgo/no procesado")

    # Exportar Excel formato detalle fila por fila
    m = re.match(r"(E\d+)_(IND|CONSO)_(\d{2})[-_](\d{4})", etiqueta)
    if m:
        base = f"{m.group(1)}_{m.group(2)}_{m.group(3)}-{m.group(4)}"
        subtitulo = f"{m.group(1)} {m.group(2)} {m.group(3)}-{m.group(4)}"
    else:
        base = re.sub(r'[\\/:*?"<>|\s]+', "_", etiqueta)
        subtitulo = etiqueta
    subtitulo += (f" — comparativo {info.get('prior_end', '?')}"
                  f" vs fuente {info.get('source_balance', '?')}"
                  f" — revisado {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    ruta = os.path.abspath(f"detalle_filas_{base}.xlsx")
    exportar_excel(
        ruta,
        "Detalle fila por fila — comparativo declarado vs. archivo fuente",
        subtitulo,
        hojas,
    )
    print(f"\nExcel generado: {ruta}")

    return 2 if total_diff else 0


async def main() -> int:
    parser = argparse.ArgumentParser(
        description="Valida comparativos de un archivo Base Notas de Workiva (no escribe nada). "
                    "Sin argumentos, pide las opciones interactivamente.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="Ej: python validar_comparativos.py   (modo interactivo)",
    )
    parser.add_argument("sociedad", nargs="?", help="Código de sociedad, ej: E110, E200")
    parser.add_argument("anio", nargs="?", help="Año base del cierre, ej: 2026")
    parser.add_argument("trimestre", nargs="?", help="Q1-Q4 (o 1-4, o mes 03/06/09/12)")
    parser.add_argument("--tipo", choices=["CONSO", "IND"], default="CONSO",
                        help="Tipo de cierre (default CONSO)")
    parser.add_argument("--id", dest="spreadsheet_id",
                        help="Spreadsheet ID directo (ignora sociedad/año/trimestre)")
    parser.add_argument("--lote", type=int, default=50, help="Hojas por lote (default 50)")
    args = parser.parse_args()

    if args.spreadsheet_id:
        ss_id, etiqueta = args.spreadsheet_id, args.spreadsheet_id
    else:
        if args.sociedad and args.anio and args.trimestre:
            sociedad, anio, tipo, trimestre = (
                args.sociedad.upper(), args.anio, args.tipo, args.trimestre
            )
            if trimestre.upper() not in MES_POR_TRIMESTRE:
                print(f"ERROR: trimestre '{trimestre}' no válido. Usa Q1-Q4, 1-4 o 03/06/09/12.")
                return 1
        else:
            sociedad, anio, tipo, trimestre = pedir_opciones()

        encontrado = await resolver_spreadsheet(sociedad, anio, trimestre, tipo)
        if not encontrado:
            return 1
        ss_id, etiqueta = encontrado

    return await validar(ss_id, etiqueta, args.lote)


if __name__ == "__main__":
    try:
        sys.exit(asyncio.run(main()))
    except KeyboardInterrupt:
        print("\nCancelado.")
        sys.exit(1)
