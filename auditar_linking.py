"""
auditar_linking.py
==================
Audita el linking de celdas en archivos Base Notas de Workiva.

Regla: si en una columna hay AL MENOS UNA celda con link (range link),
todas las celdas con valor en esa columna deberian tener link.
Las que tienen valor pero no link se reportan como SIN LINK.

Salida: Excel local  auditoria_linking_{MM}-{YYYY}.xlsx

USO:
    python auditar_linking.py
"""

import json, re, sys, time, warnings
from pathlib import Path
from collections import defaultdict

import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from urllib3.exceptions import InsecureRequestWarning
warnings.filterwarnings("ignore", category=InsecureRequestWarning)

# ---------------------------------------------------------------------------
# CREDENCIALES
# ---------------------------------------------------------------------------
CLIENT_ID     = "db2c551e-e18a-417e-8e52-d182716b8ef2"
CLIENT_SECRET = "wk_secret:oa2c:DzlUCmBQDv6raPxG09me"
WORKSPACE_ID  = "w_34913aadaa38420eabd7e4d341b78a1a"

TOKEN_URL  = "https://api.app.wdesk.com/iam/v1/oauth2/token"
WDESK_BASE = "https://api.app.wdesk.com"
VERIFY_SSL = False

SKIP_SHEETS = {
    "CP", "Bases", "Query BPC", "Query HANA AF", "Reporte en $",
    "Query - HANA - Deudores", "A.- Activos PPT",
    "B.- Patrimonio y Pasivos PPT", "C.- Estado de resultado por funcion PPT",
    "E1 Res Acumulado", "F1 Cuadraje Hoja A.- Saldo Inicial de Caja",
}

# ---------------------------------------------------------------------------
# AUTH
# ---------------------------------------------------------------------------
def get_token():
    resp = requests.post(
        TOKEN_URL,
        json={
            "grant_type":    "client_credentials",
            "client_id":     CLIENT_ID,
            "client_secret": CLIENT_SECRET,
        },
        headers={"Content-Type": "application/json"},
        timeout=30,
        verify=VERIFY_SSL,
    )
    resp.raise_for_status()
    return resp.json()["access_token"]

def get_session():
    token = get_token()
    s = requests.Session()
    s.headers.update({
        "Authorization": f"Bearer {token}",
        "Content-Type":  "application/json",
        "X-Version":     "2022-01-01",
    })
    s.verify = VERIFY_SSL
    return s

session = get_session()
_last_token = [time.time()]

def refresh_token():
    if time.time() - _last_token[0] > 480:
        ns = get_session()
        for k, v in ns.headers.items():
            session.headers[k] = v
        _last_token[0] = time.time()

# ---------------------------------------------------------------------------
# HELPERS API
# ---------------------------------------------------------------------------
def col_letter(idx):
    if idx < 26: return chr(65 + idx)
    return chr(64 + idx // 26) + chr(65 + idx % 26)

def load_all_files(mm, yyyy, tipo):
    """Carga todos los archivos Base Notas del periodo."""
    pattern = re.compile(rf"^E\d+_{tipo}_{mm}[-_]{yyyy}_Base Notas .+$")
    files = []
    url = WDESK_BASE + f"/platform/v1/files?workspaceId={WORKSPACE_ID}&limit=100"
    while url:
        r    = session.get(url, timeout=90)
        data = r.json()
        for f in data.get("data", []):
            name = f["name"]
            if pattern.match(name):
                m = re.match(r"(E\d+)", name)
                files.append({"id": f["id"], "name": name, "code": m.group(1) if m else name})
        url = data.get("@nextLink")
    return sorted(files, key=lambda x: x["code"])

def get_sheets(ss_id):
    """Retorna lista de {name, id, table_id} para cada hoja."""
    sheets = []
    url = WDESK_BASE + f"/platform/v1/spreadsheets/{ss_id}/sheets?$top=100"
    for _ in range(5):
        try:
            r    = session.get(url, timeout=90)
            data = r.json()
            for s in data.get("data", []):
                table_id = None
                if isinstance(s.get("table"), dict):
                    table_id = s["table"].get("id")
                elif isinstance(s.get("table"), str):
                    table_id = s["table"]
                sheets.append({
                    "name":     s["name"],
                    "id":       s["id"],
                    "table_id": table_id,
                })
            nxt = data.get("@nextLink")
            if not nxt:
                break
            url = nxt
        except Exception as e:
            print(f"    Error get_sheets: {e}")
            time.sleep(3)
    return sheets

def get_cell_values(ss_id, sheet_id):
    """
    Retorna dict {(row, col): {"value": ..., "is_formula": bool}}
    Indices base-0.
    """
    url = (WDESK_BASE + f"/platform/v1/spreadsheets/{ss_id}/sheets/{sheet_id}"
           + "/sheetdata?$fields=cells.calculatedValue,cells.value&$maxcellsperpage=50000")
    for _ in range(5):
        try:
            data = session.get(url, timeout=120).json()
            cells_raw = data.get("data", {}).get("cells", [])
            result = {}
            for r_idx, row in enumerate(cells_raw):
                for c_idx, cell in enumerate(row):
                    if not isinstance(cell, dict):
                        continue
                    cv  = cell.get("calculatedValue")
                    val = cell.get("value", "")
                    is_formula = str(val).startswith("=")
                    if cv is not None and str(cv).strip() not in ("", "None"):
                        result[(r_idx, c_idx)] = {
                            "calculated": cv,
                            "is_formula": is_formula,
                        }
            return result
        except Exception as e:
            print(f"    Error get_cell_values: {e}")
            time.sleep(3)
    return {}

def get_range_links(table_id):
    """
    Retorna set de (row, col) de celdas que tienen range link (source o destination).
    Usa el endpoint prototype con paginacion.
    """
    if not table_id:
        return set()
    linked = set()
    url = WDESK_BASE + f"/prototype/platform/content/tables/{table_id}/rangeLinks?$maxpagesize=500"
    while url:
        try:
            r    = session.get(url, timeout=60)
            if r.status_code == 404:
                break
            data = r.json()
            for link in data.get("data", []):
                src = link.get("source", {})
                rng = src.get("range", {})
                r1  = rng.get("startRow", 0)
                c1  = rng.get("startColumn", 0)
                r2  = rng.get("stopRow", r1)
                c2  = rng.get("stopColumn", c1)
                for ri in range(r1, r2 + 1):
                    for ci in range(c1, c2 + 1):
                        linked.add((ri, ci))
            url = data.get("@nextLink")
        except Exception as e:
            print(f"    Error get_range_links: {e}")
            break
    return linked

# ---------------------------------------------------------------------------
# AUDITORIA POR HOJA
# ---------------------------------------------------------------------------
def auditar_hoja(ss_id, sheet):
    """
    Devuelve lista de hallazgos:
    [{"celda": "C5", "valor": 123, "link": True/False}, ...]
    Solo reporta celdas SIN link en columnas donde hay AL MENOS UNA celda con link.
    """
    refresh_token()
    name     = sheet["name"]
    sheet_id = sheet["id"]
    table_id = sheet["table_id"]

    cells   = get_cell_values(ss_id, sheet_id)
    linked  = get_range_links(table_id)

    if not cells:
        return []

    # Determinar que columnas tienen al menos 1 celda linkeada
    cols_con_link = set()
    for (r, c) in linked:
        if (r, c) in cells:
            cols_con_link.add(c)

    if not cols_con_link:
        return []

    # Filas que son encabezado (primeras 8 filas con texto no numerico)
    header_rows = set()
    for (r, c), info in cells.items():
        if r < 8:
            cv = str(info["calculated"]).strip()
            try:
                float(cv.replace(",", "").replace(".", "").replace("(", "-").replace(")", ""))
            except ValueError:
                header_rows.add(r)

    hallazgos = []
    for (r, c), info in sorted(cells.items()):
        if c not in cols_con_link:
            continue
        if r in header_rows:
            continue
        tiene_link = (r, c) in linked
        if not tiene_link:
            hallazgos.append({
                "celda":      col_letter(c) + str(r + 1),
                "fila":       r + 1,
                "columna":    c + 1,
                "valor":      info["calculated"],
                "es_formula": info["is_formula"],
            })

    return hallazgos

# ---------------------------------------------------------------------------
# EXCEL DE SALIDA
# ---------------------------------------------------------------------------
VERDE  = PatternFill("solid", fgColor="C6EFCE")
ROJO   = PatternFill("solid", fgColor="FFC7CE")
AZUL   = PatternFill("solid", fgColor="BDD7EE")
NEGRITA = Font(bold=True)

def crear_excel(resultados, periodo, tipo):
    wb   = openpyxl.Workbook()
    ws   = wb.active
    ws.title = "Auditoria Linking"

    # Encabezados
    hdrs = ["Archivo", "Codigo", "Hoja", "Celda", "Fila", "Columna",
            "Valor", "Es formula", "Estado"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font      = NEGRITA
        cell.fill      = AZUL
        cell.alignment = Alignment(horizontal="center")

    row = 2
    total_sin_link = 0
    for archivo, hojas in resultados.items():
        code = re.match(r"(E\d+)", archivo)
        code = code.group(1) if code else archivo
        for hoja, hallazgos in hojas.items():
            total_sin_link += len(hallazgos)
            for h in hallazgos:
                ws.cell(row=row, column=1, value=archivo)
                ws.cell(row=row, column=2, value=code)
                ws.cell(row=row, column=3, value=hoja)
                ws.cell(row=row, column=4, value=h["celda"])
                ws.cell(row=row, column=5, value=h["fila"])
                ws.cell(row=row, column=6, value=h["columna"])
                ws.cell(row=row, column=7, value=str(h["valor"]))
                ws.cell(row=row, column=8, value="Si" if h["es_formula"] else "No")
                estado = ws.cell(row=row, column=9, value="SIN LINK")
                estado.fill = ROJO
                row += 1

    # Resumen por archivo
    ws2 = wb.create_sheet("Resumen")
    ws2.cell(1, 1, "Archivo").font     = NEGRITA
    ws2.cell(1, 2, "Codigo").font      = NEGRITA
    ws2.cell(1, 3, "Hojas auditadas").font = NEGRITA
    ws2.cell(1, 4, "Celdas sin link").font = NEGRITA
    for c in range(1, 5):
        ws2.cell(1, c).fill = AZUL

    r2 = 2
    for archivo, hojas in resultados.items():
        code = re.match(r"(E\d+)", archivo)
        code = code.group(1) if code else archivo
        sin_link = sum(len(h) for h in hojas.values())
        ws2.cell(r2, 1, archivo)
        ws2.cell(r2, 2, code)
        ws2.cell(r2, 3, len(hojas))
        c4 = ws2.cell(r2, 4, sin_link)
        if sin_link > 0:
            c4.fill = ROJO
        else:
            c4.fill = VERDE
        r2 += 1

    # Ajustar anchos
    for ws_ in [ws, ws2]:
        for col in ws_.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws_.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)

    nombre = f"auditoria_linking_{tipo}_{periodo}.xlsx"
    wb.save(nombre)
    print(f"\n  Excel guardado: {nombre}")
    print(f"  Total celdas sin link: {total_sin_link}")
    return nombre

# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    print("=" * 60)
    print("  AUDITORIA DE LINKING - WORKIVA")
    print("=" * 60)
    print()

    yyyy = input("  Ano del periodo (ej: 2026): ").strip()
    mm   = input("  Mes del periodo (ej: 03): ").strip().zfill(2)
    tipo = input("  Tipo (IND o CONSO): ").strip().upper()

    if not yyyy.isdigit() or not mm.isdigit() or not (1 <= int(mm) <= 12):
        print("  Ano o mes invalido.")
        input("\n  Enter para salir...")
        return
    if tipo not in ("IND", "CONSO"):
        print("  Tipo invalido.")
        input("\n  Enter para salir...")
        return

    periodo = f"{mm}-{yyyy}"
    print(f"\n  Buscando archivos {tipo} para {periodo}...")

    archivos = load_all_files(mm, yyyy, tipo)
    if not archivos:
        print(f"  No se encontraron archivos.")
        input("\n  Enter para salir...")
        return

    print(f"  {len(archivos)} archivos encontrados.")
    for i, a in enumerate(archivos, 1):
        print(f"    {i}. {a['name']}")

    confirm = input("\n  Procesar todos (s), elegir (e), cancelar (n)? ").strip().lower()
    if confirm == "n":
        input("\n  Enter para salir...")
        return
    if confirm == "e":
        sel = []
        for a in archivos:
            r = input(f"  Procesar {a['name']}? (s/n): ").strip().lower()
            if r == "s":
                sel.append(a)
        archivos = sel
    if not archivos:
        input("\n  Enter para salir...")
        return

    resultados = {}

    for i, arch in enumerate(archivos, 1):
        print(f"\n  [{i}/{len(archivos)}] {arch['name']}")
        hojas = get_sheets(arch["id"])
        hojas = [h for h in hojas if h["name"] not in SKIP_SHEETS]
        print(f"    {len(hojas)} hojas a auditar")

        resultados[arch["name"]] = {}
        for hoja in hojas:
            print(f"    Auditando: {hoja['name']} ...", end=" ", flush=True)
            hallazgos = auditar_hoja(arch["id"], hoja)
            print(f"{len(hallazgos)} sin link")
            if hallazgos:
                resultados[arch["name"]][hoja["name"]] = hallazgos

    crear_excel(resultados, periodo, tipo)

    print()
    print("=" * 60)
    input("  Enter para salir...")

if __name__ == "__main__":
    main()
