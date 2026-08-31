#!/usr/bin/env python3
"""
Inyecta montos desde un .xlsx en las plantillas .xlsm de DBNeT (XBRL / CMF)
sin romper las macros, los botones ni el formato.

No usa openpyxl para escribir: openpyxl descarta los dibujos (xl/drawings/*),
o sea borra los botones "Crear CSV" / "Crear/Comprimir CSV" / "Copiar Columna".
En su lugar hace cirugia directa sobre el ZIP: reescribe solo los bytes de las
celdas de monto dentro de xl/worksheets/sheetN.xml y copia todo lo demas
—vbaProject.bin incluido— byte a byte.

El calce se hace por el URI del concepto XBRL que la plantilla lleva en la
columna C (p.ej. full_ifrs-cor_2015-03-11.xsd#ifrs-full_Inventories), que es
estable entre ejercicios. Las celdas con formula (subtotales) nunca se tocan.

Uso:
    # 1. ver que conceptos y columnas tiene la plantilla
    python inyectar_montos_dbnet.py listar plantilla.xlsm
    python inyectar_montos_dbnet.py listar plantilla.xlsm --csv conceptos.csv

    # 2. inyectar los montos de tu xlsx
    python inyectar_montos_dbnet.py inyectar plantilla.xlsm montos.xlsx -o salida.xlsm
    python inyectar_montos_dbnet.py inyectar plantilla.xlsm montos.xlsx -o salida.xlsm --dry-run

    # 3. confirmar que la macro y los botones siguen intactos
    python inyectar_montos_dbnet.py verificar plantilla.xlsm salida.xlsm

El .xlsx de montos necesita una columna llave (el concepto o la etiqueta) y una
o mas columnas de valores cuyo encabezado nombre el periodo:

    concepto                                          | actual  | anterior
    full_ifrs-cor_2015-03-11.xsd#ifrs-full_Inventories| 1234567 | 1100000
"""

import argparse
import csv
import hashlib
import re
import shutil
import sys
import zipfile
from pathlib import Path

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

# Encabezados de columna de montos vistos en las plantillas DBNeT, agrupados
# por el periodo al que corresponden. Un encabezado que no este aqui igual
# sirve: se usa su propio texto normalizado como clave, asi que basta con
# nombrar la columna del .xlsx de origen igual que en la plantilla.
PERIODOS = {
    "actual": ("periodo actual", "trimestre acumulado ano actual",
               "acumulado actual"),
    "anterior": ("periodo anterior", "cierre anual anterior",
                 "trimestre acumulado ano anterior", "acumulado anterior",
                 "periodo comparativo"),
    "trim_actual": ("ultimo trimestre ano actual", "ultimo trimestre actual"),
    "trim_anterior": ("ultimo trimestre ano anterior",
                      "ultimo trimestre anterior"),
}


def clave_periodo(texto):
    """Alias canonico del encabezado, o su texto normalizado si no lo conozco."""
    n = normaliza(texto)
    if not n:
        return None
    if n in PERIODOS:
        return n
    for alias, variantes in PERIODOS.items():
        if any(n == normaliza(v) for v in variantes):
            return alias
    return n

COL_CONCEPTO = "C"   # URI del elemento de taxonomia
COL_PERIODO = "D"    # marca ACT/ANT en las plantillas dimensionales
COL_ETIQUETA = "F"   # etiqueta legible; los montos van siempre a su derecha

MARCAS_PERIODO = {"ACT": "actual", "ANT": "anterior"}


# --------------------------------------------------------------------------
# utilidades de referencias de celda
# --------------------------------------------------------------------------

def col_a_num(col):
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch.upper()) - 64)
    return n


def num_a_col(n):
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def parte_ref(ref):
    m = re.match(r"([A-Z]+)(\d+)", ref)
    return m.group(1), int(m.group(2))


def normaliza(txt):
    """Minusculas, sin acentos ni espacios repetidos: para comparar etiquetas."""
    if txt is None:
        return ""
    txt = str(txt).lower().strip()
    for a, b in zip("áéíóúüñ", "aeiouun"):
        txt = txt.replace(a, b)
    txt = re.sub(r"\[.*?\]", " ", txt)      # descarta [Número], [sinopsis], ...
    return re.sub(r"\s+", " ", txt).strip()


# --------------------------------------------------------------------------
# lectura de la plantilla
# --------------------------------------------------------------------------

def leer_shared_strings(z):
    try:
        raw = z.read("xl/sharedStrings.xml").decode("utf-8")
    except KeyError:
        return []
    out = []
    for si in re.findall(r"<si>(.*?)</si>", raw, re.S):
        txt = "".join(re.findall(r"<t[^>]*>(.*?)</t>", si, re.S))
        out.append(desescapa(txt))
    return out


def desescapa(s):
    return (s.replace("&lt;", "<").replace("&gt;", ">")
             .replace("&quot;", '"').replace("&apos;", "'")
             .replace("&amp;", "&"))


def leer_hojas(z):
    """[(nombre, parte, oculta)] en el orden del libro."""
    wb = z.read("xl/workbook.xml").decode("utf-8")
    rels = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]*)"',
                           z.read("xl/_rels/workbook.xml.rels").decode("utf-8")))
    hojas = []
    for tag in re.findall(r"<sheet [^>]*/>", wb):
        nombre = desescapa(re.search(r'name="([^"]*)"', tag).group(1))
        rid = re.search(r'r:id="(rId\d+)"', tag).group(1)
        oculta = 'state="hidden"' in tag or 'state="veryHidden"' in tag
        destino = rels[rid].lstrip("/")
        parte = destino if destino.startswith("xl/") else "xl/" + destino
        hojas.append((nombre, parte, oculta))
    return hojas


CELDA_RE = re.compile(r'<c r="([A-Z]+\d+)"([^>]*?)(?:/>|>(.*?)</c>)', re.S)
FILA_RE = re.compile(r'<row r="(\d+)"([^>]*?)(?:/>|>(.*?)</row>)', re.S)


def celdas_de_hoja(xml, shared):
    """{ref: (attrs, tipo, valor_texto, tiene_formula)}"""
    out = {}
    for ref, attrs, inner in CELDA_RE.findall(xml):
        inner = inner or ""
        tipo = (re.search(r't="(\w+)"', attrs) or [None, None])[1] \
            if 't="' in attrs else None
        m = re.search(r't="(\w+)"', attrs)
        tipo = m.group(1) if m else None
        formula = "<f" in inner
        mv = re.search(r"<v>(.*?)</v>", inner, re.S)
        val = mv.group(1) if mv else None
        if tipo == "s" and val is not None and val.isdigit():
            val = shared[int(val)] if int(val) < len(shared) else ""
        elif tipo == "inlineStr":
            val = "".join(re.findall(r"<t[^>]*>(.*?)</t>", inner, re.S))
        out[ref] = (attrs, tipo, desescapa(val) if val else val, formula)
    return out


def es_hoja_de_cuadro(celdas):
    """Firma DBNeT: toda hoja de cuadro lleva URIs de taxonomia; las
    auxiliares (listas SI/NO, Terminado/En proceso) no."""
    return any(v[2] and ".xsd#" in v[2] for v in celdas.values())


def columnas_de_monto(celdas):
    """{col: periodo} desde encabezados tipo 'Periodo Actual'.

    Solo columnas a la derecha de la de etiquetas: en las plantillas
    dimensionales esos mismos textos aparecen DENTRO de la columna F como
    titulo de bloque, y tomarlos por encabezado escribiria sobre las
    etiquetas."""
    limite = col_a_num(COL_ETIQUETA)
    candidatas, ambiguas = {}, set()
    for ref, (_, _, val, _) in celdas.items():
        if not val or ".xsd#" in val:
            continue
        col, fila = parte_ref(ref)
        if col_a_num(col) <= limite:
            continue
        clave = clave_periodo(val)
        if clave is None or clave not in PERIODOS:
            continue                       # solo encabezados de periodo conocidos
        if col in candidatas and candidatas[col][0] != clave:
            ambiguas.add(col)
        candidatas.setdefault(col, (clave, fila))

    cols, vistas = {}, {}
    for col, (clave, _fila) in sorted(candidatas.items(), key=lambda kv: col_a_num(kv[0])):
        if col in ambiguas:
            continue
        if clave in vistas:
            # la plantilla repite el mismo encabezado en dos columnas (bug de
            # DBNeT en 320000): no adivinamos, hay que pasar --col
            cols.pop(vistas[clave], None)
            cols[col] = None
            cols[vistas[clave]] = None
            continue
        vistas[clave] = col
        cols[col] = clave
    return cols


def es_dimensional(celdas):
    """True si el periodo va por fila en la columna D (ACT/ANT) en vez de por
    columna. Son las plantillas con dimensiones (CIRC1901, CLCP): cada columna
    de datos es un miembro distinto, asi que el calce necesita ademas el
    miembro y no basta con el concepto."""
    return any(
        (val or "").strip().upper() in MARCAS_PERIODO
        for ref, (_, _, val, _) in celdas.items()
        if parte_ref(ref)[0] == COL_PERIODO
    )


def analizar(ruta, override=None):
    """[(hoja, parte, [filas])] solo de hojas de cuadros."""
    with zipfile.ZipFile(ruta) as z:
        shared = leer_shared_strings(z)
        resultado = []
        for nombre, parte, oculta in leer_hojas(z):
            xml = z.read(parte).decode("utf-8")
            celdas = celdas_de_hoja(xml, shared)
            if not es_hoja_de_cuadro(celdas):
                continue
            if es_dimensional(celdas):
                resultado.append((nombre, parte, {}, [], "dimensional"))
                continue
            cols = columnas_de_monto(celdas)
            if override:
                cols = dict(override)
            filas = []
            for ref, (_, _, val, _) in celdas.items():
                col, nfila = parte_ref(ref)
                if col != COL_CONCEPTO or not val or ".xsd#" not in val:
                    continue
                etiqueta = celdas.get(f"{COL_ETIQUETA}{nfila}", (None,) * 4)[2]
                destinos = {}
                for c, periodo in cols.items():
                    if periodo is None:
                        continue
                    info = celdas.get(f"{c}{nfila}")
                    destinos[periodo] = {
                        "ref": f"{c}{nfila}",
                        "formula": bool(info and info[3]),
                        "existe": info is not None,
                    }
                filas.append({
                    "fila": nfila,
                    "concepto": val,
                    "etiqueta": (etiqueta or "").strip(),
                    "destinos": destinos,
                })
            filas.sort(key=lambda r: r["fila"])
            tipo = "columnas" if cols else "sin-montos"
            resultado.append((nombre, parte, cols, filas, tipo))
    return resultado


# --------------------------------------------------------------------------
# escritura
# --------------------------------------------------------------------------

def formatea_numero(valor):
    if isinstance(valor, bool):
        raise ValueError("valor booleano no admitido")
    f = float(valor)
    return str(int(f)) if f == int(f) else repr(f)


def escribe_celda(xml, ref, valor):
    """Reemplaza (o inserta) la celda `ref` conservando su atributo s=.
    valor None -> celda vacia. Devuelve (xml, True/False si cambio)."""
    patron = re.compile(r'<c r="%s"([^>]*?)(?:/>|>(.*?)</c>)' % re.escape(ref), re.S)
    m = patron.search(xml)

    if m:
        attrs = m.group(1)
        attrs = re.sub(r'\s*t="\w+"', "", attrs)   # el numero no lleva t=
        nuevo = (f'<c r="{ref}"{attrs}/>' if valor is None
                 else f'<c r="{ref}"{attrs}><v>{formatea_numero(valor)}</v></c>')
        if nuevo == m.group(0):
            return xml, False
        return xml[:m.start()] + nuevo + xml[m.end():], True

    if valor is None:
        return xml, False
    return _inserta_celda(xml, ref, valor), True


def _inserta_celda(xml, ref, valor):
    """Inserta una celda que no existe en el XML, en su posicion de columna."""
    col, nfila = parte_ref(ref)
    objetivo = col_a_num(col)
    nueva = f'<c r="{ref}"><v>{formatea_numero(valor)}</v></c>'

    mf = re.search(r'<row r="%d"([^>]*?)(?:/>|>(.*?)</row>)' % nfila, xml, re.S)
    if not mf:
        raise KeyError(f"la fila {nfila} no existe en la hoja; "
                       f"no se inserta para no alterar la estructura")

    attrs, inner = mf.group(1), mf.group(2) or ""
    pos = len(inner)
    for m in CELDA_RE.finditer(inner):
        if col_a_num(parte_ref(m.group(1))[0]) > objetivo:
            pos = m.start()
            break
    inner = inner[:pos] + nueva + inner[pos:]
    return (xml[:mf.start()] + f'<row r="{nfila}"{attrs}>{inner}</row>'
            + xml[mf.end():])


def fuerza_recalculo(wb_xml):
    """Excel recalcula los subtotales al abrir el archivo."""
    if "fullCalcOnLoad" in wb_xml:
        return wb_xml
    if "<calcPr" in wb_xml:
        return re.sub(r"<calcPr([^>]*?)/>", r'<calcPr\1 fullCalcOnLoad="1"/>',
                      wb_xml, count=1)
    return wb_xml.replace("</workbook>", '<calcPr fullCalcOnLoad="1"/></workbook>')


def reescribe_zip(origen, destino, cambios):
    """Copia el .xlsm entero cambiando solo las partes de `cambios`.
    Todo lo demas (vbaProject.bin, drawings, styles, media) se copia igual."""
    with zipfile.ZipFile(origen) as zin, \
         zipfile.ZipFile(destino, "w", zipfile.ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            datos = cambios.get(info.filename)
            datos = datos.encode("utf-8") if datos is not None else zin.read(info.filename)
            nuevo = zipfile.ZipInfo(info.filename, date_time=info.date_time)
            nuevo.compress_type = info.compress_type
            nuevo.external_attr = info.external_attr
            nuevo.internal_attr = info.internal_attr
            nuevo.create_system = info.create_system
            zout.writestr(nuevo, datos)


# --------------------------------------------------------------------------
# origen de datos (.xlsx)
# --------------------------------------------------------------------------

def lee_montos(ruta, hoja=None):
    """Devuelve (por_concepto, por_etiqueta) -> {llave: {periodo: valor}}"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("Falta openpyxl para leer el origen:  pip install openpyxl")

    wb = load_workbook(ruta, data_only=True, read_only=True)
    ws = wb[hoja] if hoja else wb.worksheets[0]
    filas = [list(f) for f in ws.iter_rows(values_only=True)]
    wb.close()
    if not filas:
        sys.exit(f"{ruta}: la hoja esta vacia")

    encabezado, inicio = None, 0
    for i, fila in enumerate(filas[:20]):
        textos = [normaliza(c) for c in fila]
        if any(t in ("concepto", "elemento", "etiqueta", "cuenta", "llave")
               for t in textos):
            encabezado, inicio = fila, i + 1
            break
    if encabezado is None:
        encabezado, inicio = filas[0], 1

    idx_llave, idx_valores = None, {}
    for j, celda in enumerate(encabezado):
        n = normaliza(celda)
        if idx_llave is None and n in ("concepto", "elemento", "etiqueta",
                                       "cuenta", "llave"):
            idx_llave = j
            continue
        clave = clave_periodo(celda)
        if clave:
            idx_valores[j] = clave
    if idx_llave is None:
        idx_llave = 0
    if not idx_valores:
        sys.exit("No encontre columnas de montos en el origen. Nombra sus "
                 "encabezados igual que en la plantilla ('Periodo Actual', "
                 "'Trimestre acumulado ano actual', ...) o usa los alias "
                 "cortos: " + ", ".join(sorted(PERIODOS)))

    por_concepto, por_etiqueta = {}, {}
    for fila in filas[inicio:]:
        if idx_llave >= len(fila):
            continue
        llave = fila[idx_llave]
        if llave is None or not str(llave).strip():
            continue
        valores = {}
        for j, periodo in idx_valores.items():
            if j < len(fila) and isinstance(fila[j], (int, float)) \
                    and not isinstance(fila[j], bool):
                valores[periodo] = fila[j]
        if not valores:
            continue
        llave = str(llave).strip()
        if ".xsd#" in llave:
            por_concepto[llave] = valores
        else:
            por_etiqueta[normaliza(llave)] = valores
    return por_concepto, por_etiqueta


# --------------------------------------------------------------------------
# comandos
# --------------------------------------------------------------------------

def parsea_override(valores):
    if not valores:
        return None
    fuera = {}
    for item in valores:
        if "=" not in item:
            sys.exit(f"--col espera COL=PERIODO, recibi {item!r}")
        col, periodo = item.split("=", 1)
        fuera[col.strip().upper()] = clave_periodo(periodo)
    return fuera


def cmd_listar(args):
    hojas = analizar(args.plantilla, parsea_override(args.col))
    if not hojas:
        sys.exit("No se detectaron hojas de cuadros en la plantilla.")
    filas_csv = []
    for nombre, _parte, cols, filas, tipo in hojas:
        if tipo == "dimensional":
            print(f"\n=== {nombre} — layout DIMENSIONAL (periodo en columna D)")
            print("    Cada columna de datos es un miembro de la dimension, asi que")
            print("    el concepto solo no identifica la celda. Se omite: llenala a mano.")
            continue
        cabecera = ", ".join(
            f"{c}={p}" if p else f"{c}=AMBIGUA" for c, p in
            sorted(cols.items(), key=lambda kv: col_a_num(kv[0]))) or "(ninguna)"
        print(f"\n=== {nombre} — columnas de monto: {cabecera} — "
              f"{len(filas)} conceptos")
        for f in filas:
            marcas = "".join(
                " " + p[:3].upper() + ("!" if d["formula"] else "")
                for p, d in sorted(f["destinos"].items()))
            print(f"  f{f['fila']:<4}{marcas:12} {f['etiqueta'][:58]:60}"
                  f" {f['concepto'].split('#')[-1]}")
            filas_csv.append({
                "hoja": nombre, "fila": f["fila"], "etiqueta": f["etiqueta"],
                "concepto": f["concepto"],
                **{p: ("FORMULA" if d["formula"] else d["ref"])
                   for p, d in f["destinos"].items()},
            })
    print("\n  ! = celda con formula (subtotal): nunca se escribe")
    if args.csv:
        campos = ["hoja", "fila", "etiqueta", "concepto"] + \
                 sorted({k for r in filas_csv for k in r} - {"hoja", "fila",
                                                             "etiqueta", "concepto"})
        with open(args.csv, "w", newline="", encoding="utf-8-sig") as fh:
            w = csv.DictWriter(fh, fieldnames=campos, extrasaction="ignore")
            w.writeheader()
            w.writerows(filas_csv)
        print(f"\nEscrito {args.csv} ({len(filas_csv)} filas)")


def cmd_inyectar(args):
    por_concepto, por_etiqueta = lee_montos(args.montos, args.hoja_origen)
    print(f"Origen: {len(por_concepto)} conceptos + {len(por_etiqueta)} etiquetas")

    hojas = analizar(args.plantilla, parsea_override(args.col))
    if not hojas:
        sys.exit("No se detectaron hojas de cuadros en la plantilla.")

    cambios, escritas, saltadas, sin_calce = {}, 0, [], []
    omitidas_hojas = []
    with zipfile.ZipFile(args.plantilla) as z:
        for nombre, parte, _cols, filas, tipo in hojas:
            if tipo != "columnas":
                omitidas_hojas.append((nombre, tipo))
                continue
            xml = z.read(parte).decode("utf-8")
            tocada = False
            for f in filas:
                valores = por_concepto.get(f["concepto"])
                origen = "concepto"
                if valores is None:
                    valores = por_etiqueta.get(normaliza(f["etiqueta"]))
                    origen = "etiqueta"
                if valores is None:
                    sin_calce.append((nombre, f["etiqueta"]))
                    continue
                for periodo, valor in valores.items():
                    destino = f["destinos"].get(periodo)
                    if destino is None:
                        continue
                    if destino["formula"]:
                        saltadas.append((nombre, destino["ref"], f["etiqueta"]))
                        continue
                    xml, cambio = escribe_celda(xml, destino["ref"], valor)
                    if cambio:
                        tocada = True
                        escritas += 1
                        if args.verbose:
                            print(f"  {nombre}!{destino['ref']:6} = {valor:>16,} "
                                  f"[{origen}] {f['etiqueta'][:40]}")
            if tocada:
                cambios[parte] = xml
        if cambios:
            cambios["xl/workbook.xml"] = fuerza_recalculo(
                z.read("xl/workbook.xml").decode("utf-8"))

    print(f"\nCeldas a escribir : {escritas}")
    print(f"Subtotales omitidos: {len(saltadas)}  (tienen formula)")
    print(f"Conceptos sin calce: {len(sin_calce)}")
    for nombre, tipo in omitidas_hojas:
        print(f"Hoja omitida       : {nombre}  ({tipo})")
    if sin_calce and args.verbose:
        for hoja, etiqueta in sin_calce[:40]:
            print(f"  - {hoja}: {etiqueta[:70]}")

    if args.dry_run:
        print("\n--dry-run: no se escribio ningun archivo.")
        return
    if not escritas:
        print("\nNada que escribir.")
        return

    salida = Path(args.salida)
    reescribe_zip(args.plantilla, salida, cambios)
    print(f"\nEscrito: {salida}")
    ok = verificar(args.plantilla, salida, silencioso=False)
    if not ok:
        sys.exit("La verificacion de integridad fallo.")


PARTES_CRITICAS = ("xl/vbaProject.bin",)


def verificar(original, nuevo, silencioso=False):
    """Confirma que el VBA y todo lo que no son hojas quedo byte a byte igual."""
    with zipfile.ZipFile(original) as a, zipfile.ZipFile(nuevo) as b:
        na, nb = set(a.namelist()), set(b.namelist())
        ok = True
        if na != nb:
            print(f"  FALTAN partes: {na - nb or '-'} | SOBRAN: {nb - na or '-'}")
            ok = False
        modificadas = []
        for nombre in sorted(na & nb):
            ha = hashlib.md5(a.read(nombre)).hexdigest()
            hb = hashlib.md5(b.read(nombre)).hexdigest()
            if ha != hb:
                modificadas.append(nombre)
        for critica in PARTES_CRITICAS:
            if critica in modificadas:
                print(f"  ROTO: {critica} cambio")
                ok = False
        botones_a = contar_botones(a)
        botones_b = contar_botones(nuevo and b)
        if botones_a != botones_b:
            print(f"  ROTO: botones {botones_a} -> {botones_b}")
            ok = False
        if not silencioso:
            print("\nVerificacion de integridad:")
            print(f"  partes del paquete : {len(nb)}/{len(na)}")
            print(f"  vbaProject.bin     : "
                  f"{'INTACTO' if 'xl/vbaProject.bin' not in modificadas else 'MODIFICADO'}")
            print(f"  botones con macro  : {botones_b} (original {botones_a})")
            print(f"  partes modificadas : {', '.join(modificadas) or 'ninguna'}")
            print(f"  resultado          : {'OK' if ok else 'FALLO'}")
    return ok


def contar_botones(z):
    total = 0
    for nombre in z.namelist():
        if nombre.startswith("xl/drawings/drawing") and nombre.endswith(".xml"):
            total += len(re.findall(r'macro="\[0\]![^"]+"',
                                    z.read(nombre).decode("utf-8")))
    return total


def cmd_verificar(args):
    if not verificar(args.original, args.nuevo):
        sys.exit(1)


def main():
    p = argparse.ArgumentParser(
        description="Inyecta montos en plantillas .xlsm de DBNeT sin romper las macros.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__)
    sub = p.add_subparsers(dest="comando", required=True)

    pl = sub.add_parser("listar", help="muestra conceptos y celdas de monto")
    pl.add_argument("plantilla")
    pl.add_argument("--csv", help="vuelca el mapa a un CSV")
    pl.add_argument("--col", action="append", metavar="COL=PERIODO",
                    help="fuerza una columna de montos, p.ej. --col H=actual")
    pl.set_defaults(func=cmd_listar)

    pi = sub.add_parser("inyectar", help="escribe los montos en una copia")
    pi.add_argument("plantilla")
    pi.add_argument("montos", help=".xlsx con los montos")
    pi.add_argument("-o", "--salida", required=True)
    pi.add_argument("--hoja-origen", help="hoja del .xlsx de montos")
    pi.add_argument("--col", action="append", metavar="COL=PERIODO",
                    help="fuerza una columna de montos, p.ej. --col H=actual "
                         "(necesario si 'listar' marca AMBIGUA)")
    pi.add_argument("--dry-run", action="store_true")
    pi.add_argument("-v", "--verbose", action="store_true")
    pi.set_defaults(func=cmd_inyectar)

    pv = sub.add_parser("verificar", help="compara integridad original vs nuevo")
    pv.add_argument("original")
    pv.add_argument("nuevo")
    pv.set_defaults(func=cmd_verificar)

    args = p.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
